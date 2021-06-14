import sys, os
import openpyxl
path = os.path.join(sys.path[0]) + '\\trekking1.xlsx'
print (path)

wb = openpyxl.load_workbook(path)
sheet = wb.active
rows = sheet.max_row
columns = sheet.max_column
 
# считывание в список списков строк из xlsx-файла
trekking = [[sheet.cell(row=i, column=j).value for j in range(1,columns+1)] for i in range(2,rows+1)]
# сортировка по возрастанию ( знак минус в key максимумы делает минимумами)
# в начале по второму, затем по первому столбцам
print(*[meal[0] for meal in sorted(trekking, key = lambda x: (-x[1], x[0]))], sep='\n')